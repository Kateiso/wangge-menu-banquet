import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from backend.models.menu import Menu, MenuItem

CATEGORY_ORDER = ["凉菜", "热菜", "汤羹", "主食", "甜品", "点心"]

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def generate_excel(menu: Menu, items: list[MenuItem]) -> io.BytesIO:
    """生成带公式的 Excel 菜单（普通格式）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "推荐菜单"

    # 列宽
    col_widths = [5, 22, 16, 8, 14, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 样式
    title_font = Font(name="微软雅黑", size=16, bold=True)
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cat_font = Font(name="微软雅黑", size=11, bold=True)
    cat_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    normal_font = Font(name="微软雅黑", size=10)
    total_font = Font(name="微软雅黑", size=11, bold=True)
    note_font = Font(name="微软雅黑", size=9, italic=True, color="888888")

    # 标题行
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value="旺阁渔村 · 推荐菜单")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 35

    # 信息行
    row = 2
    date_str = getattr(menu, 'date', '') or datetime.now().strftime("%Y-%m-%d")
    table_count = getattr(menu, 'table_count', 1)
    pricing_mode = getattr(menu, 'pricing_mode', 'additive')

    info_parts = [f"客户: {menu.customer_name or '贵宾'}",
                  f"人数: {menu.party_size}位",
                  f"日期: {date_str}"]
    if table_count > 1:
        info_parts.append(f"桌数: {table_count}")
    if pricing_mode == 'fixed':
        info_parts.append(f"固定价: {int(getattr(menu, 'fixed_price', 0))}元/桌")
    info = "  |  ".join(info_parts)

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value=info)
    cell.font = Font(name="微软雅黑", size=10)
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 22

    row = 3  # 空行

    # 按类别分组
    grouped: dict[str, list[MenuItem]] = {}
    for item in items:
        grouped.setdefault(item.category, []).append(item)

    subtotal_cells: list[str] = []
    has_market_price = False

    row = 4
    for cat in CATEGORY_ORDER:
        cat_items = grouped.get(cat)
        if not cat_items:
            continue

        # 类别标题
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=f"【{cat}】")
        cell.font = cat_font
        cell.fill = cat_fill
        cell.alignment = Alignment(horizontal="left")
        for c in range(1, 7):
            ws.cell(row=row, column=c).fill = cat_fill
            ws.cell(row=row, column=c).border = thin_border
        row += 1

        # 表头
        headers = ["序号", "菜名", "单价(¥)", "数量", "小计", "备注"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        row += 1

        # 菜品行
        for idx, item in enumerate(cat_items, 1):
            is_market_price = "时价" in item.price_text
            if is_market_price:
                has_market_price = True

            # 使用 adjusted_price（如有），否则用 price
            display_price = getattr(item, 'adjusted_price', 0.0)
            if not display_price or display_price <= 0:
                display_price = item.price

            ws.cell(row=row, column=1, value=idx).font = normal_font
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")

            name_text = item.dish_name
            spec_name = getattr(item, 'spec_name', '')
            if spec_name:
                name_text = f"{item.dish_name}（{spec_name}）"
            ws.cell(row=row, column=2, value=name_text).font = normal_font

            ws.cell(row=row, column=3, value=display_price).font = normal_font
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="right")

            ws.cell(row=row, column=4, value=item.quantity).font = normal_font
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

            subtotal_cell_ref = f"E{row}"
            ws.cell(row=row, column=5, value=f"=C{row}*D{row}").font = normal_font
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
            subtotal_cells.append(subtotal_cell_ref)

            note = item.reason or ""
            if is_market_price:
                note = f"{note} (时价)".strip()
            ws.cell(row=row, column=6, value=note).font = normal_font

            for c in range(1, 7):
                ws.cell(row=row, column=c).border = thin_border

            row += 1

        row += 1  # 类别间空行

    # 合计行
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="单桌合计")
    cell.font = total_font
    cell.alignment = Alignment(horizontal="right")

    if subtotal_cells:
        sum_formula = "=" + "+".join(subtotal_cells)
        ws.cell(row=row, column=5).value = sum_formula
    else:
        ws.cell(row=row, column=5, value=0)
    ws.cell(row=row, column=5).font = total_font
    ws.cell(row=row, column=5).number_format = '#,##0.00'
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")

    for c in range(1, 7):
        ws.cell(row=row, column=c).border = thin_border

    # 多桌总价
    if table_count > 1:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=f"总计（{table_count}桌）")
        cell.font = total_font
        cell.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=5, value=f"=E{row-1}*{table_count}").font = total_font
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = thin_border

    # 时价备注
    if has_market_price:
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value="* 时价菜品按参考价计算，实际以当日市价为准")
        cell.font = note_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_margin_excel(menu: Menu, items: list[MenuItem], is_admin: bool = True) -> io.BytesIO:
    """生成毛利核算表（含公式，角色控制列可见性）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "毛利核算表"

    # 列定义: A序号 B菜名 C规格 D数量 E成本 F原售价 G调整后售价 H单菜毛利 I单菜毛利率
    col_widths = [5, 20, 10, 6, 10, 10, 12, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    title_font = Font(name="微软雅黑", size=14, bold=True)
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    normal_font = Font(name="微软雅黑", size=10)
    total_font = Font(name="微软雅黑", size=11, bold=True)
    note_font = Font(name="微软雅黑", size=9, italic=True, color="888888")
    editable_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

    table_count = getattr(menu, 'table_count', 1)
    date_str = getattr(menu, 'date', '') or datetime.now().strftime("%Y-%m-%d")

    # 标题
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value="旺阁渔村 · 毛利核算表")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30

    # 信息行
    row = 2
    info = f"客户: {menu.customer_name or '贵宾'}  |  人数: {menu.party_size}位  |  桌数: {table_count}  |  日期: {date_str}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=info)
    cell.font = Font(name="微软雅黑", size=10)
    cell.alignment = Alignment(horizontal="center")

    # 表头
    row = 4
    col_headers = ["序号", "菜名", "规格", "数量", "成本(¥)", "原售价(¥)", "调整后售价(¥)", "单菜毛利(¥)", "单菜毛利率"]
    for c, h in enumerate(col_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 数据行
    data_start_row = 5
    sorted_items = sorted(items, key=lambda x: CATEGORY_ORDER.index(x.category) if x.category in CATEGORY_ORDER else 99)

    for idx, item in enumerate(sorted_items):
        r = data_start_row + idx
        ws.cell(row=r, column=1, value=idx + 1).font = normal_font
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")

        ws.cell(row=r, column=2, value=item.dish_name).font = normal_font

        spec_name = getattr(item, 'spec_name', '') or ''
        ws.cell(row=r, column=3, value=spec_name).font = normal_font

        ws.cell(row=r, column=4, value=item.quantity).font = normal_font
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")

        # E: 成本
        ws.cell(row=r, column=5, value=item.cost).font = normal_font
        ws.cell(row=r, column=5).number_format = '#,##0.00'

        # F: 原售价
        ws.cell(row=r, column=6, value=item.price).font = normal_font
        ws.cell(row=r, column=6).number_format = '#,##0.00'

        # G: 调整后售价（可编辑，高亮）
        adjusted = getattr(item, 'adjusted_price', 0.0)
        if not adjusted or adjusted <= 0:
            adjusted = item.price
        ws.cell(row=r, column=7, value=adjusted).font = normal_font
        ws.cell(row=r, column=7).number_format = '#,##0.00'
        ws.cell(row=r, column=7).fill = editable_fill

        # H: 单菜毛利 = G - E
        ws.cell(row=r, column=8, value=f"=G{r}-E{r}").font = normal_font
        ws.cell(row=r, column=8).number_format = '#,##0.00'

        # I: 单菜毛利率 = H/G
        ws.cell(row=r, column=9, value=f"=IF(G{r}>0,H{r}/G{r},0)").font = normal_font
        ws.cell(row=r, column=9).number_format = '0.0%'

        for c in range(1, 10):
            ws.cell(row=r, column=c).border = thin_border

    data_end_row = data_start_row + len(sorted_items) - 1

    # 汇总区域
    row = data_end_row + 2

    # 单桌总价
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="单桌总价").font = total_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=6, value=f"=SUMPRODUCT(G{data_start_row}:G{data_end_row},D{data_start_row}:D{data_end_row})").font = total_font
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    for c in range(1, 10):
        ws.cell(row=row, column=c).border = thin_border
    per_table_price_row = row

    # 总成本
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="单桌总成本").font = total_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=6, value=f"=SUMPRODUCT(E{data_start_row}:E{data_end_row},D{data_start_row}:D{data_end_row})").font = total_font
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    for c in range(1, 10):
        ws.cell(row=row, column=c).border = thin_border
    cost_row = row

    # 单桌毛利
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="单桌毛利").font = total_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=6, value=f"=F{per_table_price_row}-F{cost_row}").font = total_font
    ws.cell(row=row, column=6).number_format = '#,##0.00'
    for c in range(1, 10):
        ws.cell(row=row, column=c).border = thin_border
    margin_row = row

    # 毛利率
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row=row, column=1, value="整单毛利率").font = total_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=6, value=f"=IF(F{per_table_price_row}>0,F{margin_row}/F{per_table_price_row},0)").font = total_font
    ws.cell(row=row, column=6).number_format = '0.0%'
    for c in range(1, 10):
        ws.cell(row=row, column=c).border = thin_border

    # 多桌总价
    if table_count > 1:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value=f"总计（{table_count}桌）").font = total_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=6, value=f"=F{per_table_price_row}*{table_count}").font = total_font
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        for c in range(1, 10):
            ws.cell(row=row, column=c).border = thin_border

    # 说明
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1, value="* G列(调整后售价)可直接修改，修改后毛利和总价公式自动刷新").font = note_font

    # Staff 模式：隐藏成本相关列
    if not is_admin:
        ws.column_dimensions['E'].hidden = True  # 成本
        ws.column_dimensions['H'].hidden = True  # 单菜毛利
        ws.column_dimensions['I'].hidden = True  # 单菜毛利率

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
